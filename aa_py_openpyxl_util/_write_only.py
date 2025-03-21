"""
Utilities for working with write-only openpyxl workbooks.

Specifically, functions which write tables side-by-side, because in write-only mode, you can't modify a row once it
has been written, because that would mean reading it back in, which is not allowed.
"""

from __future__ import annotations

from dataclasses import dataclass
from itertools import zip_longest
from logging import getLogger
from typing import Optional, Any, Sequence, Generator, List, Iterable, Callable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell, Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet._write_only import WriteOnlyWorksheet
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import TableStyleInfo

from ._list_objects import define_list_object
from ._written_tables_types import (
    WrittenTables,
    WrittenTablesInSheet,
)

logger = getLogger(__name__)


@dataclass
class FormattedCell:
    """
    Custom class to hold cell data separate from a sheet.

    This is required because:

        - openpyxl does not allow creating a WriteOnlyCell with formatting if it's not attached to a sheet.
        - openpyxl's `ArrayFormula` class needs a `ref` value, which can only be derived when we know the cell position.
    """

    value: Any
    """
    The cell value. If it's a string starting with `=`, openpyxl interprets it as a formula.
    """

    number_format: Optional[str] = None
    """
    The cell's number format. Optional.
    """

    array: bool = False
    """
    Whether this is an array formula. This sets `t="array"` in the `f` tag in XML.

    FIXME: This actually makes a CSE formula, which is not quite what we want, but it works for some use cases.
    See https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1898
    """

    def check(self) -> FormattedCell:
        """
        Check a cell for potential errors before writing it to a sheet.

        If writing the given cell would cause the workbook to be invalid / corrupted / unusable, raise a user-friendly
        error message here instead of letting Excel crash later.

        Returns:
            `self`, for convenience.

        Raises:
            ValueError: If the cell would cause problems.
        """
        if isinstance(self.value, str) and self.value.startswith("="):
            # This is a formula. Check that it's not longer than 8192 characters.
            formula = self.value[1:]
            if len(formula) > 8192:
                raise ValueError(
                    f"Formula is too long: {len(formula)} characters. The maximum is 8192.\n{formula}"
                )

        return self

    def create_openpyxl_cell(
        self,
        sheet: WriteOnlyWorksheet,
        ref: str,
    ) -> Cell:
        """

        Args:
            sheet:
                The sheet into which the cell will be written.
            ref:
                The value of the `ref` attribute for the `f` tag in XML. Required when `self.array==True`. For array
                formulas with scalar results, this should refer to the cell containing the formula. For array formulas
                with multi-celled results, this should refer to the entire range of cells that will contain the results.

        Returns:
            An openpyxl `WriteOnlyCell` instance.
        """
        value = (
            ArrayFormula(
                ref=ref,
                text=self.value,
            )
            if self.array
            else self.value
        )

        cell: Cell = WriteOnlyCell(ws=sheet, value=value)

        if self.number_format:
            # noinspection PyUnresolvedReferences,PyDunderSlots
            cell.number_format = self.number_format

        return cell


default_table_style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)


class TableInfo:
    """
    Info about a table which is to be written to a sheet using `openpyxl`.
    """

    name: str
    """
    The table name
    """

    column_names: Sequence[str]
    """
    The column names.
    """

    n_rows: int
    """
    The number of rows in the table.
    """

    get_cell: Callable[[int, int], FormattedCell]
    """
    A function that returns a cell for a given row and column index. The indices are 0-based.
    """

    pre_rows: Sequence[Sequence[FormattedCell]]
    """
    Rows to write outside the table, above the header, but below the name and description.
    This may be wider or narrower as the table if required.
    """

    style: Optional[TableStyleInfo]
    """
    The table style
    """

    description: str
    """
    A table description to write below the table name.
    """

    def __init__(
        self,
        *,
        name: str,
        column_names: Sequence[str],
        n_rows: int | None = None,
        get_cell: Callable[[int, int], FormattedCell] | None = None,
        rows: Sequence[Sequence[FormattedCell]] | None = None,
        pre_rows: Sequence[Sequence[FormattedCell]] | None = None,
        style: Optional[TableStyleInfo] | None = None,
        description: str | None = None,
    ):
        self.name = name
        self.column_names = column_names
        self.pre_rows = pre_rows or []
        self.style = style or default_table_style
        self.description = description or ""

        if rows is None:
            if n_rows is None or get_cell is None:
                raise ValueError(
                    "Either `rows` or (`n_rows` and `get_cell`) must be provided."
                )
            self.n_rows = n_rows
            self.get_cell = get_cell
        else:
            if n_rows is not None or get_cell is not None:
                raise ValueError(
                    "`rows` and (`n_rows` and `get_cell`) are mutually exclusive."
                )
            self.n_rows = len(rows)
            self.get_cell = lambda i_row, i_col: rows[i_row][i_col]  # type: ignore[index]

    @property
    def width(self) -> int:
        """
        The width of the table. This is not necessarily the same as the number of columns, because `pre_rows` is
        allowed to be wider or narrower than the rest of the table.
        """
        return max(
            [
                len(self.column_names),
                *(len(r) for r in self.pre_rows),
            ]
        )

    @property
    def rows(self) -> Generator[Generator[FormattedCell, None, None], None, None]:
        for i_row in range(self.n_rows):
            yield (
                self.get_cell(i_row, i_col) for i_col in range(len(self.column_names))
            )


def write_tables_side_by_side_over_multiple_sheets(
    *,
    book: Workbook,
    base_sheet_name: str,
    tables: Sequence[TableInfo],
    row_margin: int,
    col_margin: int,
    col_margin_width: int | None = None,
    write_captions: bool,
    write_pre_rows: bool,
    max_sheet_width: int,
) -> WrittenTables:
    """
    Create one or more sheets containing one or more tables, stacked horizontally.

    Like `write_tables_side_by_side`, but automatically uses multiple sheets as necessary.

    Args:
        book: A write-only workbook in which to create the sheet and table.
        base_sheet_name:
            The name of the first sheet. If there are too many tables to fit on one sheet, a counter is appended,
            starting at `1`. E.g.: `Tables`, `Tables1`, `Tables2`, etc.
        tables: A sequence of table info objects.
        row_margin: The number of empty rows to leave above each table.
        col_margin: The number of empty columns to leave to the left of each table.
        col_margin_width: The width of the margin columns. If None, the column width is left at the default.
        write_captions: Whether to write the table name and description above the table. This shifts the table down.
        write_pre_rows: Whether to write the pre_rows (below the name and description, but above the table header).
        max_sheet_width:
            The maximum number of columns to write to a single sheet. If the tables are too wide, they will be split
            across multiple sheets. The maximum sheet width in Excel from 2007 is 16384 columns. Before 2007 it was 256
            columns. See https://support.microsoft.com/en-us/office/use-excel-with-earlier-versions-of-excel-2fd9ffcb-6fce-485b-85af-fecfd651a5ac

    Returns: A dictionary with:
        - Keys: The sheet names.
        - Values: A dictionary with:
            - Keys: The table names.
            - Values: A tuple with:
                - The co-ordinates of the top-left cell of the table (e.g. `(2,3)` which means cell C2)
                - The openpyxl table object.
    """
    result: WrittenTables = {}
    for i, tables_in_sheet in enumerate(
        distribute_tables_over_multiple_sheets(
            tables=tables,
            max_sheet_width=max_sheet_width,
            left_margin=col_margin,
            gutter=col_margin,
        )
    ):
        sheet_name = base_sheet_name if i == 0 else f"{base_sheet_name}{i}"
        result[sheet_name] = write_tables_side_by_side(
            book=book,
            sheet_name=sheet_name,
            tables=tables_in_sheet,
            row_margin=row_margin,
            col_margin=col_margin,
            col_margin_width=col_margin_width,
            write_captions=write_captions,
            write_pre_rows=write_pre_rows,
        )

    return result


def write_tables_side_by_side(
    *,
    book: Workbook,
    sheet_name: str,
    tables: Sequence[TableInfo],
    row_margin: int,
    col_margin: int,
    col_margin_width: int | None = None,
    write_captions: bool,
    write_pre_rows: bool,
) -> WrittenTablesInSheet:
    """
    Create a new sheet containing one or more tables, stacked horizontally.

    If the tables don't all fit into one sheet, an error will be raised. If you need this situation to be handled
    transparently, use `write_tables_side_by_side_over_multiple_sheets`.

    See https://openpyxl.readthedocs.io/en/stable/worksheet_tables.html#creating-a-table
    See https://openpyxl.readthedocs.io/en/stable/worksheet_tables.html#manually-adding-column-headings

    Args:
        book: A write-only workbook in which to create the sheet and table.
        sheet_name: The name of the new sheet.
        tables: A sequence of table info objects.
        row_margin: The number of empty rows to leave above each table.
        col_margin: The number of empty columns to leave to the left of each table.
        col_margin_width: The width of the margin columns. If None, the column width is left at the default.
        write_captions: Whether to write the table name and description above the table. This shifts the table down.
        write_pre_rows: Whether to write the pre_rows (below the name and description, but above the table header).

    Returns: A dictionary with:
        - Keys: The table names.
        - Values: A tuple with:
            - The co-ordinates of the top-left cell of the table (e.g. `(2,3)` which means cell C2)
            - The openpyxl table object.
    """
    sheet: WriteOnlyWorksheet = book.create_sheet(title=sheet_name)

    # Write rows
    for i_row, row in enumerate(
        stack_table_rows_side_by_side(
            tables=tables,
            row_margin=row_margin,
            col_margin=col_margin,
            write_captions=write_captions,
            write_pre_rows=write_pre_rows,
        ),
        start=1,
    ):
        sheet.append(
            (
                cell.check().create_openpyxl_cell(
                    sheet=sheet,
                    ref=f"{get_column_letter(i_col)}{i_row}",
                )
                for i_col, cell in enumerate(row, start=1)
            )
        )

    # Define ListObjects
    results: WrittenTablesInSheet = {}

    if not len(tables):
        return results

    first_row = (
        1
        + row_margin
        + (2 if write_captions else 0)
        + (max(len(t.pre_rows) for t in tables) if write_pre_rows else 0)
    )
    first_column = 1 + col_margin
    for t in tables:
        width = t.width
        if width < 1:
            raise ValueError(f"Can't create table '{t.name}' with zero columns.")

        coords = first_row, first_column
        lo = define_list_object(
            sheet=sheet,
            first_column=first_column,
            first_row=first_row,
            name=t.name,
            column_names=t.column_names,
            n_data_rows=t.n_rows,
            style=t.style,
        )
        results[t.name] = (coords, lo)

        first_column += col_margin + width

    if col_margin and col_margin_width:
        # Set the width of the gutter columns.
        for name, ((r, c), table) in results.items():
            if c > 1:
                sheet.column_dimensions[get_column_letter(c - 1)].width = (
                    col_margin_width
                )

    return results


def distribute_tables_over_multiple_sheets(
    *,
    tables: Iterable[TableInfo],
    max_sheet_width: int,
    left_margin: int,
    gutter: int,
) -> List[List[TableInfo]]:
    """
    Distribute tables over multiple sheets, so that the total width of each sheet is within bounds.

    Args:
        tables: The table to distribute.
        max_sheet_width: The maximum number of columns in each sheet.
        left_margin: The number of columns to leave open left of the first table.
        gutter: The number of columns to keep open between tables.

    Returns:
        A list of lists of tables. Each inner list represents a sheet.
    """
    result: List[List[TableInfo]] = []
    current_column = 1 + left_margin  # 1-indexed, like the Excel, i.e., column A is 1.
    current_sheet = 0
    for table in tables:
        if len(result) < 1:
            result.append([])

        table_width = table.width
        if len(result[current_sheet]) != 0:
            current_column += gutter

        if current_column + table_width - 1 > max_sheet_width:
            # This table does not fit in the current sheet.

            if len(result[current_sheet]) == 0:
                # This is the first table in the sheet.
                # This table will never fit.
                raise ValueError(
                    f"Table `{table.name}` is too wide to fit in a sheet with maximum width {max_sheet_width} "
                    f"and a left margin of {left_margin} columns."
                )

            # Move to the next sheet.
            current_sheet += 1
            current_column = 1 + left_margin
            result.append([])

        # Add the table to the current sheet.
        result[current_sheet].append(table)
        current_column += table_width

    return result


def stack_table_rows_side_by_side(
    tables: Sequence[TableInfo],
    row_margin: int,
    col_margin: int,
    write_captions: bool,
    write_pre_rows: bool,
) -> Generator[List[FormattedCell], None, None]:
    """
    Iterate over the cells of multiple tables simultaneously, and yield one row at a time,

    Args:
        tables: A sequence of table info objects.
        row_margin: The number of empty rows to leave above each table.
        col_margin: The number of empty columns to leave to the left of each table.
        write_captions: Whether to write the table name and description above the table.
        write_pre_rows: Whether to write the pre_rows (below the name and description, but above the table header).

    Returns:
        A generator that yields one row at a time, in such a way that a sheet can be written from this data,
        top to bottom, without ever going back to a previous row.
    """
    if row_margin < 0:
        raise ValueError("Row margin must be a positive integer.")
    if col_margin < 0:
        raise ValueError("Column margin must be a positive integer.")

    def table_name_row() -> Generator[FormattedCell, None, None]:
        for t in tables:
            yield from [FormattedCell(None)] * col_margin
            yield FormattedCell(t.name)
            yield from [FormattedCell(None)] * (t.width - 1)

    def table_description_row() -> Generator[FormattedCell, None, None]:
        for t in tables:
            yield from [FormattedCell(None)] * col_margin
            yield FormattedCell(t.description)
            yield from [FormattedCell(None)] * (t.width - 1)

    def header_row() -> Generator[FormattedCell, None, None]:
        for t in tables:
            yield from [FormattedCell(None)] * col_margin
            for c in t.column_names:
                yield FormattedCell(c)
            yield from [FormattedCell(None)] * (t.width - len(t.column_names))

    widths = [t.width for t in tables]

    def row(
        data: Iterable[Optional[Iterable[FormattedCell]]],
    ) -> Generator[FormattedCell, None, None]:
        for w, d in zip(widths, data):
            yield from [FormattedCell(None)] * col_margin

            if d is None:
                yield from [FormattedCell(None)] * w
            else:
                d_len = 0
                for cell in d:
                    yield cell
                    d_len += 1
                yield from [FormattedCell(None)] * (w - d_len)

    for _ in range(row_margin):
        yield []

    if write_captions:
        yield list(table_name_row())
        yield list(table_description_row())

    if write_pre_rows:
        for pre_row_data in zip_longest(*(t.pre_rows for t in tables), fillvalue=None):
            yield list(row(pre_row_data))

    yield list(header_row())

    n_data_rows = 0
    for row_data in zip_longest(*(t.rows for t in tables), fillvalue=None):
        yield list(row(row_data))
        n_data_rows += 1

    if n_data_rows < 1:
        # Tables are not allowed to have zero rows. Add an empty row.
        yield []
