"""
Utilities for working with read-only openpyxl workbooks.
"""
from contextlib import contextmanager
from pathlib import Path
from typing import Generator, Dict

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


@contextmanager
def safe_load_workbook(
    *,
    path: Path,
    read_only: bool,
    data_only: bool,
) -> Generator[Workbook, None, None]:
    """
    Open a workbook with openpyxl. Make sure the file handle is closed afterwards.

    This is a context manager.

    See Also:
        https://stackoverflow.com/questions/31416842/openpyxl-does-not-close-excel-workbook-in-read-only-mode

    Args:
        path: The path to the workbook on the disk.
        read_only: See https://openpyxl.readthedocs.io/en/stable/optimized.html?highlight=read_only#read-only-mode
        data_only: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.workbook.workbook.html?highlight=data_only#openpyxl.workbook.workbook.Workbook.data_only

    Yields:
        The workbook.
    """
    book: Workbook = load_workbook(
        filename=path,
        read_only=read_only,
        data_only=data_only,
    )
    try:
        yield book
    finally:
        book.close()


@contextmanager
def changed_builtin_number_formats(
    formats: Dict[int, str]
) -> Generator[None, None, None]:
    """
    Temporarily change one or more of `openpyxl`'s builtin number formats.

    This is necessary because `openpyxl` does not take regional settings into account.
    """
    from openpyxl.styles.numbers import BUILTIN_FORMATS

    original_values = {key: BUILTIN_FORMATS[key] for key in formats}
    BUILTIN_FORMATS.update(formats)

    try:
        yield
    finally:
        BUILTIN_FORMATS.update(original_values)
