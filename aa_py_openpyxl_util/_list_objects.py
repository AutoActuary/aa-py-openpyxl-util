from __future__ import annotations

import warnings
from typing import Sequence, Optional

from openpyxl.utils import get_column_letter
from openpyxl.worksheet._write_only import WriteOnlyWorksheet
from openpyxl.worksheet.table import TableStyleInfo, Table


def define_list_object(
    sheet: WriteOnlyWorksheet,
    first_column: int,
    first_row: int,
    name: str,
    column_names: Sequence[str],
    n_data_rows: int,
    style: Optional[TableStyleInfo],
) -> Table:
    last_column = first_column - 1 + len(column_names)
    last_row = first_row + max(n_data_rows, 1)

    table = Table(
        displayName=name,
        ref=f"{get_column_letter(first_column)}{first_row}:{get_column_letter(last_column)}{last_row}",
    )
    # noinspection PyProtectedMember
    table._initialise_columns()
    for column, value in zip(table.tableColumns, column_names):
        column.name = value

    if style:
        table.tableStyleInfo = style

    with warnings.catch_warnings():
        # See https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1760
        warnings.simplefilter(action="ignore", category=UserWarning)
        sheet.add_table(table)

    return table
