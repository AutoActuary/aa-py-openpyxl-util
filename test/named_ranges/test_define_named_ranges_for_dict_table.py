import unittest

from openpyxl import Workbook

from aa_py_openpyxl_util import define_named_ranges_for_dict_table


class TestDefineNamedRangesForDictTable(unittest.TestCase):
    def test_define_named_ranges(self) -> None:
        book = Workbook()
        sheet = book.active
        sheet.title = "Sheet1"
        keys = ["key1", "key2", "key3"]

        define_named_ranges_for_dict_table(
            book=book,
            sheet_name="Sheet1",
            first_table_row=1,
            first_table_col=1,
            keys=keys,
            workbook_scope=True,
        )

        for i, key in enumerate(keys):
            defined_name = book.defined_names[key]
            self.assertIsNotNone(defined_name)
            self.assertEqual(
                defined_name.attr_text,
                f"'Sheet1'!$B${i+2}:$B${i+2}",
            )

    def test_skip_none_keys(self) -> None:
        book = Workbook()
        sheet = book.active
        sheet.title = "Sheet1"
        keys = ["key1", None, "key3"]

        define_named_ranges_for_dict_table(
            book=book,
            sheet_name="Sheet1",
            first_table_row=1,
            first_table_col=1,
            keys=keys,
            workbook_scope=True,
        )

        self.assertEqual(book.defined_names["key1"].attr_text, "'Sheet1'!$B$2:$B$2")
        self.assertIsNone(book.defined_names.get("key2"))
        self.assertEqual(book.defined_names["key3"].attr_text, "'Sheet1'!$B$4:$B$4")

    def test_sheet_scope_named_ranges(self) -> None:
        book = Workbook()
        sheet = book.active
        sheet.title = "Sheet1"
        keys = ["key1", "key2", "key3"]

        define_named_ranges_for_dict_table(
            book=book,
            sheet_name="Sheet1",
            first_table_row=1,
            first_table_col=1,
            keys=keys,
            workbook_scope=False,
        )

        for i, key in enumerate(keys):
            defined_name = sheet.defined_names[key]
            self.assertIsNotNone(defined_name)
            self.assertEqual(
                defined_name.attr_text,
                f"'Sheet1'!$B${i+2}:$B${i+2}",
            )
