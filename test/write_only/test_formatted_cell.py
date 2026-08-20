import sys
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl.styles import Font
from openpyxl.workbook import Workbook

from aa_py_openpyxl_util import FormattedCell


class TestFormattedCell(unittest.TestCase):
    def test_short_formula(self) -> None:
        fc = FormattedCell(value="=SUM(A1:A10)", number_format="0.00")
        checked = fc.check()
        self.assertIs(fc, checked)

    def test_maximum_length_formula(self) -> None:
        # Create a formula of the maximum length that Excel can support.
        s = '"abcdefghijklmnopqrstuvwxyz"'
        concat_args = ", ".join([s] * 136)
        concat = f"_xlfn.CONCAT({concat_args})"
        f = f'{concat}&{concat}&"abcd"'
        self.assertEqual(8192, len(f))

        fc = FormattedCell(value=f"={f}", number_format="0.00")
        checked = fc.check()
        self.assertIs(fc, checked)

        # Check manually with Excel to verify that this is indeed a valid workbook:
        # from openpyxl.workbook import Workbook
        # book = Workbook(write_only=True)
        # ws = book.create_sheet()
        # ws.append([fc.create_openpyxl_cell(ws, "A1")])
        # book.save("max.xlsx")

    def test_too_long_formula(self) -> None:
        # Create a formula that is too long for Excel.
        s = '"abcdefghijklmnopqrstuvwxyz"'
        concat_args = ", ".join([s] * 136)
        concat = f"_xlfn.CONCAT({concat_args})"
        f = f'{concat}&{concat}&"abcde"'
        self.assertEqual(8193, len(f))

        fc = FormattedCell(value=f"={f}", number_format="0.00")
        with self.assertRaises(ValueError):
            fc.check()

        # Check manually with Excel to verify that this is indeed a corrupted workbook:
        # from openpyxl.workbook import Workbook
        # book = Workbook(write_only=True)
        # ws = book.create_sheet()
        # ws.append([fc.create_openpyxl_cell(ws, "A1")])
        # book.save("too_long.xlsx")

    def test_hyperlinks(self) -> None:
        cells = [
            # Hyperlink formulas:
            FormattedCell(
                value='=HYPERLINK("https://www.autoactuary.com")',
                hyperlink=False,
            ),
            FormattedCell(
                value='=HYPERLINK("https://www.autoactuary.com")',
                hyperlink=True,
            ),
            # Hyperlink values:
            FormattedCell(
                value="https://www.autoactuary.com",
                hyperlink=False,
            ),
            FormattedCell(
                value="https://www.autoactuary.com",
                hyperlink=True,
            ),
            # Combine with custom fonts:
            FormattedCell(
                value='=HYPERLINK("https://www.autoactuary.com")',
                hyperlink=True,
                font=Font(bold=True),
            ),
            FormattedCell(
                value="https://www.autoactuary.com",
                hyperlink=True,
                font=Font(name="Consolas"),
            ),
        ]

        book = Workbook(write_only=True)
        ws = book.create_sheet()
        for i, cell in enumerate(cells):
            ws.append([cell.create_openpyxl_cell(sheet=ws, ref=f"A{i+1}")])

        with TemporaryDirectory() as tmp_dir_str:
            tmp_dir = Path(tmp_dir_str)
            book_path = tmp_dir / "hyperlinks.xlsx"
            book.save(book_path)

            if sys.platform != "win32":
                return

            # Check using Excel to see what the hyperlinks look like.
            try:
                # noinspection unused-imports
                from aa_py_xl import excel
            except ImportError:
                return

            with excel(
                path=book_path,
                save=False,
                quiet=True,
                events=False,
                close_excel=True,
                close_book=True,
                must_exist=True,
                read_only=True,
            ) as xl_book:
                from xlwings.constants import ThemeColor, UnderlineStyle

                sheet = xl_book.sheets[0]
                hyperlink_refs = {"A2", "A4", "A5", "A6"}

                for row in range(1, 7):
                    ref = f"A{row}"
                    font = sheet.range(ref).api.Font
                    with self.subTest(ref=ref):
                        if ref in hyperlink_refs:
                            self.assertEqual(
                                ThemeColor.xlThemeColorHyperlink,
                                font.ThemeColor,
                            )
                            self.assertEqual(
                                UnderlineStyle.xlUnderlineStyleSingle,
                                font.Underline,
                            )
                        else:
                            self.assertEqual(0, font.Color)
                            self.assertEqual(
                                UnderlineStyle.xlUnderlineStyleNone,
                                font.Underline,
                            )

                self.assertTrue(sheet.range("A5").api.Font.Bold)
                self.assertEqual("Consolas", sheet.range("A6").api.Font.Name)
