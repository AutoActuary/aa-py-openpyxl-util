import unittest
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Callable

from openpyxl import Workbook
from openpyxl.worksheet.table import TableStyleInfo

from aa_py_openpyxl_util import (
    safe_load_workbook,
    TableInfo,
    write_tables_side_by_side,
    FormattedCell,
    find_table,
    get_cell_values,
    write_tables_side_by_side_over_multiple_sheets,
)


class TestWriteTablesSideBySide(unittest.TestCase):
    def test_empty(self) -> None:
        def write(book: Workbook) -> None:
            results = write_tables_side_by_side(
                book=book,
                sheet_name="Sheet1",
                tables=[
                    TableInfo(
                        name="Table1",
                        column_names=["a", "b", "c"],
                        rows=[],
                    ),
                    TableInfo(
                        name="Table2",
                        column_names=["d", "e"],
                        rows=[],
                    ),
                ],
                row_margin=1,
                col_margin=1,
                write_captions=False,
                write_pre_rows=False,
            )

            self.assertEqual([(2, 2), (2, 6)], [c for c, t in results.values()])

        def test(book: Workbook) -> None:
            table1_sheet, table1_range = find_table(book=book, name="Table1")
            self.assertEqual("B2:D3", table1_range)
            self.assertEqual(
                [["a", "b", "c"], [None, None, None]],
                get_cell_values(table1_sheet[table1_range]),
            )

            table2_sheet, table2_range = find_table(book=book, name="Table2")
            self.assertEqual("F2:G3", table2_range)
            self.assertEqual(
                [["d", "e"], [None, None]],
                get_cell_values(table2_sheet[table2_range]),
            )

        test_helper(write, test)

    def test_same_length(self) -> None:
        def write(book: Workbook) -> None:
            results = write_tables_side_by_side(
                book=book,
                sheet_name="Sheet1",
                tables=[
                    TableInfo(
                        name="Table1",
                        column_names=["a", "b", "c"],
                        rows=[
                            [FormattedCell(1), FormattedCell(2), FormattedCell(3)],
                            [FormattedCell(4), FormattedCell(5), FormattedCell(6)],
                        ],
                    ),
                    TableInfo(
                        name="Table2",
                        column_names=["d", "e"],
                        rows=[
                            [FormattedCell(7), FormattedCell(8)],
                            [FormattedCell(9), FormattedCell(0)],
                        ],
                    ),
                ],
                row_margin=1,
                col_margin=1,
                write_captions=False,
                write_pre_rows=False,
            )

            self.assertEqual([(2, 2), (2, 6)], [c for c, t in results.values()])

        def test(book: Workbook) -> None:
            table1_sheet, table1_range = find_table(book=book, name="Table1")
            self.assertEqual("B2:D4", table1_range)
            self.assertEqual(
                [["a", "b", "c"], [1, 2, 3], [4, 5, 6]],
                get_cell_values(table1_sheet[table1_range]),
            )

            table2_sheet, table2_range = find_table(book=book, name="Table2")
            self.assertEqual("F2:G4", table2_range)
            self.assertEqual(
                [["d", "e"], [7, 8], [9, 0]],
                get_cell_values(table2_sheet[table2_range]),
            )

        test_helper(write, test)

    def test_different_lengths(self) -> None:
        def write(book: Workbook) -> None:
            results = write_tables_side_by_side(
                book=book,
                sheet_name="Sheet1",
                tables=[
                    TableInfo(
                        name="Table1",
                        column_names=["a", "b", "c"],
                        rows=[
                            [FormattedCell(1), FormattedCell(2), FormattedCell(3)],
                            [FormattedCell(2), FormattedCell(3), FormattedCell(4)],
                            [FormattedCell(3), FormattedCell(4), FormattedCell(5)],
                        ],
                    ),
                    TableInfo(
                        name="Table2",
                        column_names=["d", "e"],
                        rows=[],
                    ),
                    TableInfo(
                        name="Table3",
                        column_names=["f", "g"],
                        rows=[
                            [FormattedCell(7), FormattedCell(8)],
                            [FormattedCell(9), FormattedCell(0)],
                        ],
                    ),
                ],
                row_margin=1,
                col_margin=1,
                write_captions=False,
                write_pre_rows=False,
            )

            self.assertEqual([(2, 2), (2, 6), (2, 9)], [c for c, t in results.values()])

        def test(book: Workbook) -> None:
            table1_sheet, table1_range = find_table(book=book, name="Table1")
            self.assertEqual("B2:D5", table1_range)
            self.assertEqual(
                [["a", "b", "c"], [1, 2, 3], [2, 3, 4], [3, 4, 5]],
                get_cell_values(table1_sheet[table1_range]),
            )

            table2_sheet, table2_range = find_table(book=book, name="Table2")
            self.assertEqual("F2:G3", table2_range)
            self.assertEqual(
                [["d", "e"], [None, None]],
                get_cell_values(table2_sheet[table2_range]),
            )

            table3_sheet, table3_range = find_table(book=book, name="Table3")
            self.assertEqual("I2:J4", table3_range)
            self.assertEqual(
                [["f", "g"], [7, 8], [9, 0]],
                get_cell_values(table3_sheet[table3_range]),
            )

        test_helper(write, test)

    def test_captions(self) -> None:
        def write(book: Workbook) -> None:
            write_tables_side_by_side(
                book=book,
                sheet_name="Sheet1",
                tables=[
                    TableInfo(
                        name="Table1",
                        column_names=["a", "b", "c"],
                        rows=[
                            [FormattedCell(1), FormattedCell(2), FormattedCell(3)],
                            [FormattedCell(2), FormattedCell(3), FormattedCell(4)],
                            [FormattedCell(3), FormattedCell(4), FormattedCell(5)],
                        ],
                        description="First table",
                    ),
                ],
                row_margin=1,
                col_margin=1,
                write_captions=True,
                write_pre_rows=False,
            )

        def test(book: Workbook) -> None:
            table1_sheet, table1_range = find_table(book=book, name="Table1")
            self.assertEqual("B4:D7", table1_range)
            self.assertEqual(
                [["a", "b", "c"], [1, 2, 3], [2, 3, 4], [3, 4, 5]],
                get_cell_values(table1_sheet[table1_range]),
            )
            self.assertEqual("Table1", table1_sheet["B2"].value)
            self.assertEqual("First table", table1_sheet["B3"].value)

        test_helper(write, test)

    def test_pre_rows(self) -> None:
        def write(book: Workbook) -> None:
            write_tables_side_by_side(
                book=book,
                sheet_name="Sheet1",
                tables=[
                    TableInfo(
                        name="Table1",
                        column_names=["a", "b", "c"],
                        rows=[
                            [FormattedCell(1), FormattedCell(2), FormattedCell(3)],
                            [FormattedCell(2), FormattedCell(3), FormattedCell(4)],
                            [FormattedCell(3), FormattedCell(4), FormattedCell(5)],
                        ],
                        pre_rows=[
                            [
                                FormattedCell("Pre A"),
                                FormattedCell("Pre B"),
                                FormattedCell("Pre C"),
                            ],
                        ],
                    ),
                ],
                row_margin=1,
                col_margin=1,
                write_captions=False,
                write_pre_rows=True,
            )

        def test(book: Workbook) -> None:
            table1_sheet, table1_range = find_table(book=book, name="Table1")
            self.assertEqual("B3:D6", table1_range)
            self.assertEqual(
                [["a", "b", "c"], [1, 2, 3], [2, 3, 4], [3, 4, 5]],
                get_cell_values(table1_sheet[table1_range]),
            )
            self.assertEqual("Pre A", table1_sheet["B2"].value)
            self.assertEqual("Pre B", table1_sheet["C2"].value)
            self.assertEqual("Pre C", table1_sheet["D2"].value)

        test_helper(write, test)

    def test_values(self) -> None:
        def write(book: Workbook) -> None:
            write_tables_side_by_side(
                book=book,
                sheet_name="Table1",
                tables=[
                    TableInfo(
                        name="Table1",
                        column_names=["a", "b", "c"],
                        rows=list(
                            [
                                [FormattedCell(1), FormattedCell(2), FormattedCell(4)],
                                [FormattedCell(2), FormattedCell(4), FormattedCell(8)],
                                [FormattedCell(3), FormattedCell(6), FormattedCell(12)],
                                [FormattedCell(4), FormattedCell(8), FormattedCell(16)],
                            ]
                        ),
                        style=(
                            TableStyleInfo(
                                name="TableStyleMedium2",
                                showFirstColumn=False,
                                showLastColumn=False,
                                showRowStripes=True,
                                showColumnStripes=False,
                            )
                        ),
                    )
                ],
                row_margin=5,
                col_margin=2,
                write_captions=False,
                write_pre_rows=False,
            )

        def test(book: Workbook) -> None:
            table_sheet, table_range = find_table(book=book, name="Table1")
            self.assertEqual(
                [["a", "b", "c"], [1, 2, 4], [2, 4, 8], [3, 6, 12], [4, 8, 16]],
                get_cell_values(table_sheet[table_range]),
            )

        test_helper(write, test)

    def test_formulas(self) -> None:
        def write(book: Workbook) -> None:
            write_tables_side_by_side(
                book=book,
                sheet_name="Table1",
                tables=[
                    TableInfo(
                        name="Table1",
                        column_names=["a", "b", "c"],
                        rows=list(
                            [
                                [
                                    FormattedCell(1),
                                    FormattedCell(
                                        "=Table1[[#This Row],[a]]*Table1[[#This Row],[a]]"
                                    ),
                                    FormattedCell(
                                        "=Table1[[#This Row],[a]]+Table1[[#This Row],[b]]"
                                    ),
                                ],
                                [
                                    FormattedCell(2),
                                    FormattedCell(
                                        "=Table1[[#This Row],[a]]*Table1[[#This Row],[a]]"
                                    ),
                                    FormattedCell(
                                        "=Table1[[#This Row],[a]]+Table1[[#This Row],[b]]"
                                    ),
                                ],
                                [
                                    FormattedCell(3),
                                    FormattedCell(
                                        "=Table1[[#This Row],[a]]*Table1[[#This Row],[a]]"
                                    ),
                                    FormattedCell(
                                        "=Table1[[#This Row],[a]]+Table1[[#This Row],[b]]"
                                    ),
                                ],
                                [
                                    FormattedCell(4),
                                    FormattedCell(
                                        "=Table1[[#This Row],[a]]*Table1[[#This Row],[a]]"
                                    ),
                                    FormattedCell(
                                        "=Table1[[#This Row],[a]]+Table1[[#This Row],[b]]"
                                    ),
                                ],
                            ]
                        ),
                        style=(
                            TableStyleInfo(
                                name="TableStyleMedium2",
                                showFirstColumn=False,
                                showLastColumn=False,
                                showRowStripes=True,
                                showColumnStripes=False,
                            )
                        ),
                    )
                ],
                row_margin=5,
                col_margin=2,
                write_captions=False,
                write_pre_rows=False,
            )

        def test(book: Workbook) -> None:
            table_sheet, table_range = find_table(book=book, name="Table1")
            self.assertEqual(
                [
                    ["a", "b", "c"],
                    [
                        1,
                        "=Table1[[#This Row],[a]]*Table1[[#This Row],[a]]",
                        "=Table1[[#This Row],[a]]+Table1[[#This Row],[b]]",
                    ],
                    [
                        2,
                        "=Table1[[#This Row],[a]]*Table1[[#This Row],[a]]",
                        "=Table1[[#This Row],[a]]+Table1[[#This Row],[b]]",
                    ],
                    [
                        3,
                        "=Table1[[#This Row],[a]]*Table1[[#This Row],[a]]",
                        "=Table1[[#This Row],[a]]+Table1[[#This Row],[b]]",
                    ],
                    [
                        4,
                        "=Table1[[#This Row],[a]]*Table1[[#This Row],[a]]",
                        "=Table1[[#This Row],[a]]+Table1[[#This Row],[b]]",
                    ],
                ],
                # Because we are testing with `data_only=False`, this will return formulas.
                # Also, `openpyxl` does not run formulas, and we don't want to depend on `xlwings` here.
                get_cell_values(table_sheet[table_range]),
            )

        test_helper(write, test)

    def test_number_format(self) -> None:
        def write(book: Workbook) -> None:
            write_tables_side_by_side(
                book=book,
                sheet_name="Table1",
                tables=[
                    TableInfo(
                        name="Table1",
                        column_names=["a"],
                        rows=list(
                            [
                                [FormattedCell(value=5, number_format="General")],
                                [FormattedCell(value=5, number_format="0")],
                                [FormattedCell(value=5, number_format="0.00")],
                                [FormattedCell(value=5, number_format="0%")],
                                [FormattedCell(value=5, number_format="yyyy/mm/dd")],
                            ]
                        ),
                        style=(
                            TableStyleInfo(
                                name="TableStyleMedium2",
                                showFirstColumn=False,
                                showLastColumn=False,
                                showRowStripes=True,
                                showColumnStripes=False,
                            )
                        ),
                    )
                ],
                row_margin=5,
                col_margin=2,
                write_captions=False,
                write_pre_rows=False,
            )

        def test(book: Workbook) -> None:
            table_sheet, table_range = find_table(book=book, name="Table1")
            self.assertEqual(
                [["a"], [5], [5], [5], [5], [datetime(1900, 1, 5, 0, 0)]],
                get_cell_values(table_sheet[table_range]),
            )

        test_helper(write, test)


class TestWriteTablesSideBySideOverMultipleSheets(unittest.TestCase):
    def test_two_sheets(self) -> None:
        table1 = TableInfo(
            name="Table1",
            column_names=["a", "b"],
            rows=list([[FormattedCell(1), FormattedCell(2)]]),
        )
        table2 = TableInfo(
            name="Table2",
            column_names=["a", "b"],
            rows=list([[FormattedCell(1), FormattedCell(2)]]),
        )

        def write(book: Workbook) -> None:
            write_tables_side_by_side_over_multiple_sheets(
                book=book,
                base_sheet_name="Tables",
                tables=[table1, table2],
                row_margin=5,
                col_margin=2,
                write_captions=False,
                write_pre_rows=False,
                max_sheet_width=5,
            )

        def test(book: Workbook) -> None:
            table1_sheet, table1_range = find_table(book=book, name="Table1")
            self.assertEqual(
                [["a", "b"], [1, 2]], get_cell_values(table1_sheet[table1_range])
            )
            self.assertEqual("Tables", table1_sheet.title)

            table2_sheet, table2_range = find_table(book=book, name="Table2")
            self.assertEqual(
                [["a", "b"], [1, 2]], get_cell_values(table2_sheet[table2_range])
            )
            self.assertEqual("Tables1", table2_sheet.title)

        test_helper(write, test)

    def test_one_sheet(self) -> None:
        table1 = TableInfo(
            name="Table1",
            column_names=["a", "b"],
            rows=list([[FormattedCell(1), FormattedCell(2)]]),
        )
        table2 = TableInfo(
            name="Table2",
            column_names=["a", "b"],
            rows=list([[FormattedCell(1), FormattedCell(2)]]),
        )

        def write(book: Workbook) -> None:
            write_tables_side_by_side_over_multiple_sheets(
                book=book,
                base_sheet_name="Tables",
                tables=[table1, table2],
                row_margin=5,
                col_margin=2,
                write_captions=False,
                write_pre_rows=False,
                max_sheet_width=10,
            )

        def test(book: Workbook) -> None:
            table1_sheet, table1_range = find_table(book=book, name="Table1")
            self.assertEqual(
                [["a", "b"], [1, 2]], get_cell_values(table1_sheet[table1_range])
            )
            self.assertEqual("Tables", table1_sheet.title)

            table2_sheet, table2_range = find_table(book=book, name="Table2")
            self.assertEqual(
                [["a", "b"], [1, 2]], get_cell_values(table2_sheet[table2_range])
            )
            self.assertEqual("Tables", table2_sheet.title)

        test_helper(write, test)


def test_helper(
    write: Callable[[Workbook], None],
    test: Callable[[Workbook], None],
) -> None:
    book = Workbook(write_only=True)
    write(book)

    with TemporaryDirectory() as tmp_dir:
        path = Path(tmp_dir, "test.xlsx")
        book.save(path)

        with safe_load_workbook(
            path=path,
            read_only=False,
            data_only=False,
        ) as book:
            test(book)


if __name__ == "__main__":
    unittest.main(
        failfast=True,
    )
