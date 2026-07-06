import unittest

from openpyxl import Workbook

from aa_py_openpyxl_util import (
    FormattedCell,
    TableInfo,
    define_named_ranges_for_dict_table,
    write_tables_side_by_side,
    write_tables_side_by_side_over_multiple_sheets,
)
from aa_py_openpyxl_util._excel_names import (
    DuplicateExcelNameError,
    InvalidExcelNameError,
    InvalidExcelSheetTitleError,
)


class TestWriteOnlyExcelNameValidation(unittest.TestCase):
    def test_invalid_table_name_raises_before_sheet_is_created(self) -> None:
        book = Workbook(write_only=True)

        with self.assertRaises(InvalidExcelNameError):
            write_tables_side_by_side(
                book=book,
                sheet_name="Sheet1",
                tables=[
                    TableInfo(
                        name="A1",
                        column_names=["a"],
                        rows=[[FormattedCell(1)]],
                    )
                ],
                row_margin=1,
                col_margin=1,
                write_captions=False,
                write_pre_rows=False,
            )

        self.assertEqual([], book.worksheets)

    def test_invalid_table_name_in_multi_sheet_write_raises_before_any_sheet_is_created(
        self,
    ) -> None:
        book = Workbook(write_only=True)

        with self.assertRaises(InvalidExcelNameError):
            write_tables_side_by_side_over_multiple_sheets(
                book=book,
                base_sheet_name="Tables",
                tables=[
                    TableInfo(
                        name="Table1",
                        column_names=["a"],
                        rows=[[FormattedCell(1)]],
                    ),
                    TableInfo(
                        name="R1C1",
                        column_names=["a"],
                        rows=[[FormattedCell(1)]],
                    ),
                ],
                row_margin=1,
                col_margin=1,
                write_captions=False,
                write_pre_rows=False,
                max_sheet_width=2,
            )

        self.assertEqual([], book.worksheets)

    def test_duplicate_table_name_raises_before_sheet_is_created(self) -> None:
        book = Workbook(write_only=True)

        with self.assertRaises(DuplicateExcelNameError):
            write_tables_side_by_side(
                book=book,
                sheet_name="Sheet1",
                tables=[
                    TableInfo(
                        name="Table1",
                        column_names=["a"],
                        rows=[[FormattedCell(1)]],
                    ),
                    TableInfo(
                        name="table1",
                        column_names=["b"],
                        rows=[[FormattedCell(2)]],
                    ),
                ],
                row_margin=1,
                col_margin=1,
                write_captions=False,
                write_pre_rows=False,
            )

        self.assertEqual([], book.worksheets)

    def test_invalid_sheet_title_raises_before_sheet_is_created(self) -> None:
        book = Workbook(write_only=True)

        with self.assertRaises(InvalidExcelSheetTitleError):
            write_tables_side_by_side(
                book=book,
                sheet_name="A" * 32,
                tables=[],
                row_margin=1,
                col_margin=1,
                write_captions=False,
                write_pre_rows=False,
            )

        self.assertEqual([], book.worksheets)


class TestNamedRangeExcelNameValidation(unittest.TestCase):
    def test_invalid_defined_name_raises_before_name_is_added(self) -> None:
        book = Workbook()
        sheet = book.active
        sheet.title = "Sheet1"

        with self.assertRaises(InvalidExcelNameError):
            define_named_ranges_for_dict_table(
                book=book,
                sheet_name="Sheet1",
                first_table_row=1,
                first_table_col=1,
                keys=["A1"],
                workbook_scope=True,
            )

        self.assertEqual([], list(book.defined_names.keys()))

    def test_duplicate_defined_name_raises_before_names_are_added(self) -> None:
        book = Workbook()
        sheet = book.active
        sheet.title = "Sheet1"

        with self.assertRaises(DuplicateExcelNameError):
            define_named_ranges_for_dict_table(
                book=book,
                sheet_name="Sheet1",
                first_table_row=1,
                first_table_col=1,
                keys=["Name1", "name1"],
                workbook_scope=True,
            )

        self.assertEqual([], list(book.defined_names.keys()))


if __name__ == "__main__":
    unittest.main(failfast=True)
