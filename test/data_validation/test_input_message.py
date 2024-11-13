import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Callable

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from aa_py_openpyxl_util import safe_load_workbook, set_data_validation_input_message


class TestSetDataValidationInputMessage(unittest.TestCase):
    def test_1(self) -> None:
        def write(book: Workbook) -> None:
            sheet: Worksheet = book.create_sheet(title="Sheet1")

            set_data_validation_input_message(
                worksheet=sheet,
                ranges=["A1", "B2"],
                title="foo",
                input_message="bar",
            )

        def test(book: Workbook) -> None:
            sheet: Worksheet = book["Sheet1"]
            data_validation = sheet.data_validations.dataValidation[0]
            self.assertEqual(data_validation.promptTitle, "foo")
            self.assertEqual(data_validation.prompt, "bar")
            self.assertEqual({r.coord for r in data_validation.sqref}, {"A1", "B2"})

        test_helper(write, test, False)


def test_helper(
    write: Callable[[Workbook], None],
    test: Callable[[Workbook], None],
    write_only: bool,
) -> None:
    book = Workbook(write_only=write_only)
    write(book)

    with TemporaryDirectory() as tmp_dir:
        path = Path(tmp_dir, "test.xlsx")
        book.save(path)

        with safe_load_workbook(
            path=path,
            read_only=False,
            data_only=False,
        ) as book:
            test(book)
